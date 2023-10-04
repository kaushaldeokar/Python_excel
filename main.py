import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import json
workbook = openpyxl.Workbook()
# header input
excel_header = [{"transporter_name": "RAJ CARRYING CARGO PVT LTD (AHMEDABAD)"},{"table_title":"CASH BOOK A/c. From : 01/04/2023 To 31/03/2024"}]


# Sample JSON data (replace this with your own JSON data)
table_data = [
    {"Vou No.": "", "Account Head":"OPENING BALANCE"," Cheque No.": 30, "Debit": 567095,"Credit":20, "Balance":567095,"":"Dr"},
    {"Vou No.": "CR-1", "Account Head":"SERVICE CHARGE"," Cheque No.": "", "Debit": 645,"Credit":20, "Balance":10,"":"Dr"},
    {"Vou No.": "", "Account Head":"DELI STATEMENT NO: 306"," Cheque No.": "", "Debit": "","Credit":"", "Balance":"","":""},
    {"Vou No.": "CR-1", "Account Head":"Daywise Total"," Cheque No.": 30, "Debit": 123456,"Credit":12234, "Balance":"","":""},

]



# Create a new Excel workbook and add a worksheet
sheet = workbook.active



# Heading
sheet.merge_cells('A1:G1')
cell = sheet.cell(row=1, column=1)
cell.value = excel_header[0]['transporter_name']
cell.alignment = Alignment(horizontal='center', vertical='center')
# fontStyle = Font(size = 15)
fontStyle = Font(name='Arial',size=13,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
cell.font = cell.font = fontStyle

# table title
sheet.merge_cells('A2:G2')
cell = sheet.cell(row=2, column=1)
cell.value = excel_header[1]['table_title']
cell.alignment = Alignment(horizontal='center', vertical='center')
# fontStyle = Font(size = 15)
fontStyle = Font(name='Arial',size=13,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
cell.font = cell.font = fontStyle


# Write the headers to the first row
headers = list(table_data[0].keys())
for col_num, header in enumerate(headers, 1):
   cell = sheet.cell(row=3, column=col_num, value=header)
   cell.alignment = Alignment(horizontal='center', vertical='center')
   fontStyle = Font(name='Arial', size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False,color='FF000000')
   cell.font = cell.font = fontStyle

# Write the data from the JSON to the worksheet
for row_num, data in enumerate(table_data, 4):
    for col_num, key in enumerate(headers, 1):
        sheet.cell(row=row_num, column=col_num, value=data[key])



# Save the workbook to a file
workbook.save('output1.xlsx')


print('Excel sheet generated successfully!')
